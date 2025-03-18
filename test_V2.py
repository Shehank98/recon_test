import streamlit as st
import pandas as pd
from io import BytesIO
import json
from fuzzywuzzy import fuzz
from datetime import time
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import re

# Utility Functions
def preprocess_time_format(time_value):
 """
 Preprocess time format to remove milliseconds or extra segments.
 Keeps only hh:mm:ss.
 """
 try:
     if isinstance(time_value, str):
         # Use regex to extract hh:mm:ss format
         match = re.match(r'^(\d{2}:\d{2}:\d{2})', time_value)
         if match:
             return match.group(1)  # Return cleaned time format
         else:
             raise ValueError(f"Invalid time format: {time_value}")
     return time_value  # Return unchanged if not a string
 except Exception as e:
     st.warning(f"Error preprocessing time '{time_value}': {e}. Returning '00:00:00'.")
     return '00:00:00'

def convert_duration(duration, duration_format, handle_nan):
 try:
     # Handle NaN values
     if pd.isna(duration) and handle_nan:
         return 0  # Replace NaN with 0 seconds
     
     # Handle hh:mm:ss format
     if duration_format == "hh:mm:ss":
         h, m, s = map(int, str(duration).strip().split(':'))
         return h * 3600 + m * 60 + s
     
     # Handle seconds format
     elif duration_format == "seconds":
         if isinstance(duration, (int, float)):
             return int(duration)  # Convert float to integer and return as seconds
         elif isinstance(duration, str) and duration.replace('.', '', 1).isdigit():
             return int(float(duration))  # Handle strings like "30.0"

     # Unsupported format
     raise ValueError(f"Unsupported duration format: {duration_format}")
 except ValueError as e:
     st.warning(f"Invalid duration format '{duration}': {e}. Returning 0 seconds.")
     return 0

def convert_time(time_value, remove_milliseconds):
 try:
     if isinstance(time_value, str):
         # Remove milliseconds or extra segments if specified in the configuration
         if remove_milliseconds:
             time_segments = time_value.split(':')[:3]  # Keep only hours, minutes, seconds
             time_value = ':'.join(time_segments)

         return pd.to_datetime(time_value, format='%H:%M:%S', errors='coerce').strftime('%H:%M:%S')
     elif isinstance(time_value, pd.Timestamp):
         return time_value.strftime('%H:%M:%S')
     elif isinstance(time_value, (pd.Timedelta, time)):
         return str(time_value)
     else:
         raise ValueError("Unexpected time format.")
 except Exception as e:
     st.warning(f"Error converting time '{time_value}': {e}. Returning '00:00:00'.")
     return '00:00:00'

def load_channel_config(channel):
 """Load channel-specific configuration from JSON file."""
 try:
     # Normalize the channel name to match the JSON file naming convention
     normalized_channel = channel.strip().replace(" ", "_").lower()
     config_file_path = f"configs/{normalized_channel}.json"
     
     st.write(f"Looking for configuration file: {config_file_path}")  # Debugging tip
     
     with open(config_file_path, "r") as config_file:
         return json.load(config_file)
 except FileNotFoundError:
     st.error(f"No configuration found for channel '{channel}'. Please ensure the configuration file exists.")
     return None

def transform_tc_data(tc_df, config):
 """Transform TC data based on channel-specific configuration."""
 try:
     # Step 1: Rename columns if specified in the configuration
     if "rename_columns" in config and config["rename_columns"]:
         tc_df = tc_df.rename(columns=config["rename_columns"])

     # Debugging: Inspect renamed columns
     st.write("Columns after renaming:", tc_df.columns)

     # Step 2: Validate required columns
     required_columns = [
         config["date_column"],
         config["duration_column"],
         config["time_column"],
         config["advt_theme_column"],
         config["program_column"]
     ]
     missing_columns = [col for col in required_columns if col not in tc_df.columns]
     if missing_columns:
         raise KeyError(f"Missing required columns in TC sheet: {missing_columns}")

     # Step 3: Apply transformations based on the configuration
     tc_df[config["date_column"]] = pd.to_datetime(tc_df[config["date_column"]], dayfirst=True)
     tc_df['Dd'] = tc_df[config["date_column"]].dt.day
     tc_df['Mn'] = tc_df[config["date_column"]].dt.month
     tc_df['Yr'] = tc_df[config["date_column"]].dt.year

     # Convert duration dynamically
     tc_df['Duration'] = tc_df[config["duration_column"]].apply(
         lambda x: convert_duration(x, config["duration_format"], config["handle_nan_duration"])
     )

     # Preprocess Advt_time column to remove milliseconds or extra segments
     tc_df[config["time_column"]] = tc_df[config["time_column"]].apply(preprocess_time_format)

     # Convert time dynamically
     if config["additional_transformations"]["convert_time_format"]:
         tc_df['Advt_time'] = tc_df[config["time_column"]].apply(
             lambda x: convert_time(x, config["remove_milliseconds"])
         )

     # Debugging: Inspect Advt_time column after conversion
     st.write("Transformed Advt_time values:", tc_df["Advt_time"].head())

     if config["additional_transformations"]["normalize_case"]:
         tc_df['Advt_Theme'] = tc_df[config["advt_theme_column"]].str.lower()
         tc_df['Program'] = tc_df[config["program_column"]].str.lower()
     else:
         tc_df['Advt_Theme'] = tc_df[config["advt_theme_column"]]
         tc_df['Program'] = tc_df[config["program_column"]]

     return tc_df[['Advt_Theme', 'Program', 'Advt_time', 'Dd', 'Mn', 'Yr', 'Duration']]
 except KeyError as e:
     st.error(str(e))
     return pd.DataFrame()

def filter_by_channel(df, channel, config):
 """Filter LMRB data by selected channel and rename columns if specified."""
 # Step 1: Rename columns if specified in the configuration
 if "rename_columns" in config and config["rename_columns"]:
     df = df.rename(columns=config["rename_columns"])

 # Step 2: Filter by channel
 return df[df["Channel"] == channel]

def remove_duplicates(current_df, previous_df):
 """Remove duplicates between current and previous month's data."""
 duplicate_columns = ["Advt_Theme", "Channel", "Program", "Advt_time", "Dd", "Mn", "Yr"]
 merged_df = pd.merge(
     current_df,
     previous_df,
     on=duplicate_columns,
     how="left",
     indicator=True
 )
 cleaned_df = merged_df[merged_df["_merge"] == "left_only"].drop(columns=["_merge"])
 duplicate_count = len(merged_df[merged_df["_merge"] == "both"])
 return cleaned_df, duplicate_count

def match_tc_with_lmrb(tc_df, lmrb_df):
 """Match TC data with LMRB data based on multiple criteria."""
 if tc_df.empty or lmrb_df.empty:
     st.error("One or both DataFrames are empty.")
     return pd.DataFrame(), []

 # Convert time columns to timedelta for comparison
 tc_df['advt_time_timedelta'] = pd.to_timedelta(tc_df['Advt_time'])
 lmrb_df['advt_time_timedelta'] = pd.to_timedelta(lmrb_df['Advt_time'])

 matched_rows = []
 matched_indices = []

 # Iterate through TC rows
 for _, tc_row in tc_df.iterrows():
     # Filter LMRB rows by matching date (day, month, year)
     date_matched_lmrb = lmrb_df[
         (lmrb_df['Dd'] == tc_row['Dd']) &
         (lmrb_df['Mn'] == tc_row['Mn']) &
         (lmrb_df['Yr'] == tc_row['Yr'])
     ]

     # Skip if no rows match the date
     if date_matched_lmrb.empty:
         continue

     # Compute similarities and find the best match
     date_matched_lmrb['time_difference'] = abs(
         (tc_row['advt_time_timedelta'] - date_matched_lmrb['advt_time_timedelta']).dt.total_seconds()
     )
     date_matched_lmrb = date_matched_lmrb[date_matched_lmrb['time_difference'] <= 50]  # Filter within 15 seconds

     if date_matched_lmrb.empty:
         continue

     # Initialize similarity scores
     date_matched_lmrb['day_similarity'] = 20  # Full marks for day match
     date_matched_lmrb['month_similarity'] = 20  # Full marks for month match
     date_matched_lmrb['year_similarity'] = 10  # Full marks for year match
     date_matched_lmrb['time_similarity'] = date_matched_lmrb['time_difference'].apply(
         lambda x: 20 - (x / 15 * 20) if x <= 15 else 0
     )
     date_matched_lmrb['duration_similarity'] = date_matched_lmrb['Dur_x'].apply(
         lambda x: 10 - (abs(tc_row['Duration'] - x) / max(tc_row['Duration'], x) * 10) if max(tc_row['Duration'], x) > 0 else 10
     )

     # Calculate fuzzy similarity scores only if values are not blank
     date_matched_lmrb['program_similarity'] = 0
     date_matched_lmrb['theme_similarity'] = 0

     if pd.notna(tc_row['Program']) and tc_row['Program'].strip() != "":
         date_matched_lmrb['program_similarity'] = date_matched_lmrb['Program'].apply(
             lambda x: fuzz.token_set_ratio(tc_row['Program'], x) / 100 * 10
         )
     if pd.notna(tc_row['Advt_Theme']) and tc_row['Advt_Theme'].strip() != "":
         date_matched_lmrb['theme_similarity'] = date_matched_lmrb['Advt_Theme'].apply(
             lambda x: fuzz.token_set_ratio(tc_row['Advt_Theme'], x) / 100 * 10
         )

     # Total similarity score
     date_matched_lmrb['total_similarity'] = (
         date_matched_lmrb['day_similarity'] +
         date_matched_lmrb['month_similarity'] +
         date_matched_lmrb['year_similarity'] +
         date_matched_lmrb['time_similarity'] +
         date_matched_lmrb['duration_similarity'] +
         date_matched_lmrb['program_similarity'] +
         date_matched_lmrb['theme_similarity']
     )

     # Find the best match above the threshold
     best_match = date_matched_lmrb.loc[date_matched_lmrb['total_similarity'].idxmax()]
     if best_match['total_similarity'] >= 50:  # Threshold for match
         matched_rows.append(best_match)
         matched_indices.append(best_match.name)  # Store the index of the matched row

 # Return matched rows as a DataFrame and their indices
 return pd.DataFrame(matched_rows), matched_indices

def highlight_matched_rows_excel(lmrb_df, matched_indices, highlight_color):
 """Highlight matched rows in the LMRB DataFrame when exporting to Excel."""
 # Add a 'Matched' column to indicate matched rows
 lmrb_df['Matched'] = False  # Initialize all rows as not matched
 lmrb_df.loc[matched_indices, 'Matched'] = True  # Mark matched rows as True

 # Create an Excel workbook
 wb = Workbook()
 ws = wb.active

 # Write headers
 for col_num, column_name in enumerate(lmrb_df.columns, start=1):
     ws.cell(row=1, column=col_num, value=column_name)

 # Write data with conditional formatting
 for row_num, row_data in enumerate(lmrb_df.itertuples(index=False), start=2):
     for col_num, cell_value in enumerate(row_data, start=1):
         cell = ws.cell(row=row_num, column=col_num, value=cell_value)
         if col_num == len(row_data):  # Check the 'Matched' column
             if row_data[-1]:  # If 'Matched' is True
                 for cell_to_highlight in ws[row_num]:
                     cell_to_highlight.fill = PatternFill(start_color=highlight_color.lstrip("#"), end_color=highlight_color.lstrip("#"), fill_type="solid")

 return wb

def main():
# Set page configuration
    st.set_page_config(page_title="LMRB Data Processor", layout="wide")

    # Sidebar for file uploads
    st.sidebar.title("Upload Files")
    uploaded_lmrb_file = st.sidebar.file_uploader("Upload LMRB Data Sheet", type=["xlsx", "xls", "csv"])
    uploaded_previous_file = st.sidebar.file_uploader("Upload Previous Month's Processed Sheet", type=["xlsx", "xls", "csv"])
    uploaded_tc_file = st.sidebar.file_uploader("Upload TC Sheet", type=["xlsx", "xls", "csv"])

    # Main title
    st.title("LMRB Data Processor")
    st.write("Process LMRB data efficiently by uploading sheets, selecting channels, and performing transformations.")

    if uploaded_lmrb_file:
        try:
            # Progress bar for file processing
            with st.spinner("Processing LMRB Data Sheet..."):
                # Read the uploaded file
                if uploaded_lmrb_file.name.endswith(".xlsx"):
                    lmrb_df = pd.read_excel(uploaded_lmrb_file, engine='openpyxl')
                elif uploaded_lmrb_file.name.endswith(".xls"):
                    lmrb_df = pd.read_excel(uploaded_lmrb_file, engine='xlrd')
                else:
                    lmrb_df = pd.read_csv(uploaded_lmrb_file)

            # Step 2: Select Channel
            if "Channel" in lmrb_df.columns:
                channel = st.selectbox("Select Channel", options=lmrb_df["Channel"].unique())
                
                # Load channel-specific configuration
                channel_config = load_channel_config(channel)
                if not channel_config:
                    st.error("Failed to load channel-specific configuration.")
                    return
                
                # Filter LMRB data and rename columns dynamically
                filtered_df = filter_by_channel(lmrb_df, channel, channel_config)
                st.success(f"Filtered Data for Channel: {channel}")
                st.dataframe(filtered_df)

                if uploaded_previous_file:
                    with st.spinner("Processing Previous Month's Sheet..."):
                        if uploaded_previous_file.name.endswith(".xlsx"):
                            previous_df = pd.read_excel(uploaded_previous_file, engine='openpyxl')
                        elif uploaded_previous_file.name.endswith(".xls"):
                            previous_df = pd.read_excel(uploaded_previous_file, engine='xlrd')
                        else:
                            previous_df = pd.read_csv(uploaded_previous_file)

                    # Remove duplicates between current and previous month's data
                    cleaned_df, duplicate_count = remove_duplicates(filtered_df, previous_df)
                    st.write(f"Duplicate Count: {duplicate_count}")
                    st.success("Cleaned Data:")
                    st.dataframe(cleaned_df)

                    # Group by Advt_Theme and get counts
                    grouped_advt_theme = cleaned_df.groupby('Advt_Theme').size().reset_index(name='Count')
                    st.success("Advertisement Theme Counts:")
                    st.dataframe(grouped_advt_theme)

                    # Buttons for downloading cleaned data and advertisement theme counts
                    col1, col2 = st.columns(2)
                    with col1:
                        output_cleaned = BytesIO()
                        cleaned_df.to_excel(output_cleaned, index=False, engine='openpyxl')
                        output_cleaned.seek(0)
                        st.download_button(
                            label="Download Cleaned Sheet",
                            data=output_cleaned,
                            file_name="cleaned_sheet.xlsx",
                            key="cleaned_sheet_download",
                            help="Download the cleaned LMRB data sheet."
                        )
                    with col2:
                        output_grouped = BytesIO()
                        grouped_advt_theme.to_excel(output_grouped, index=False, engine='openpyxl')
                        output_grouped.seek(0)
                        st.download_button(
                            label="Download Advertisement Theme Counts",
                            data=output_grouped,
                            file_name="advt_theme_counts.xlsx",
                            key="advt_theme_counts_download",
                            help="Download the grouped advertisement theme counts."
                        )

                    if uploaded_tc_file:
                        with st.spinner("Processing TC Sheet..."):
                            if uploaded_tc_file.name.endswith(".xlsx"):
                                tc_df = pd.read_excel(uploaded_tc_file, engine='openpyxl')
                            elif uploaded_tc_file.name.endswith(".xls"):
                                tc_df = pd.read_excel(uploaded_tc_file, engine='xlrd')
                            else:
                                tc_df = pd.read_csv(uploaded_tc_file)

                        # Transform TC data based on channel-specific configuration
                        transformed_tc_df = transform_tc_data(tc_df, channel_config)
                        st.success("Transformed TC Data:")
                        st.dataframe(transformed_tc_df)

                        # Match TC data with cleaned LMRB data
                        matched_df, matched_indices = match_tc_with_lmrb(transformed_tc_df, cleaned_df)
                        st.success("Matched Data:")
                        st.dataframe(matched_df)

                        total_matched_count = len(matched_df)
                        st.write(f"Total Matched Data: {total_matched_count}")

                        # Choose highlight color
                        highlight_color = st.color_picker("Choose Highlight Color", "#FFFF00")

                        # Highlight matched rows in cleaned LMRB and export to Excel
                        highlighted_workbook = highlight_matched_rows_excel(cleaned_df, matched_indices, highlight_color)

                        # Save the workbook to a BytesIO object
                        output_highlighted = BytesIO()
                        highlighted_workbook.save(output_highlighted)
                        output_highlighted.seek(0)

                        # Buttons for downloading highlighted LMRB data and matched data
                        col3, col4 = st.columns(2)
                        with col3:
                            st.download_button(
                                label="Download Highlighted LMRB Sheet",
                                data=output_highlighted,
                                file_name="highlighted_lmrb.xlsx",
                                key="highlighted_lmrb_download",
                                help="Download the highlighted LMRB data sheet."
                            )
                        with col4:
                            matched_output = BytesIO()
                            matched_df.to_excel(matched_output, index=False, engine='openpyxl')
                            matched_output.seek(0)
                            st.download_button(
                                label="Download Matched Data",
                                data=matched_output,
                                file_name="matched_data.xlsx",
                                key="matched_data_download",
                                help="Download the matched data sheet."
                            )
            else:
                st.error("The uploaded LMRB file does not contain a 'Channel' column.")
        except Exception as e:
            st.error(f"Error processing the uploaded LMRB file: {e}")
    else:
        st.info("Please upload an LMRB data sheet to begin.")

if __name__ == "__main__":
    main()

