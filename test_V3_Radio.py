import streamlit as st
import pandas as pd
from io import BytesIO
import json
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from fuzzywuzzy import fuzz

# Utility Functions
def load_channel_config(channel):
 """Load channel-specific configuration from JSON file."""
 try:
     # Normalize the channel name to match the JSON file naming convention
     normalized_channel = channel.strip().replace(" ", "_").lower()
     config_file_path = f"configs/{normalized_channel}.json"

     st.write(f"Looking for configuration file: {config_file_path}")  # Debugging tip

     with open(config_file_path, "r") as config_file:
         return json.load(config_file)
 except FileNotFoundError:
     st.error(f"No configuration found for channel '{channel}'. Please ensure the configuration file exists.")
     return None

def transform_tc_data(tc_df, config):
  """Transform TC data based on channel-specific configuration."""
  try:
      # Step 1: Rename columns if specified in the configuration
      if "rename_columns" in config and config["rename_columns"]:
          tc_df = tc_df.rename(columns=config["rename_columns"])

      # Debugging: Inspect renamed columns
      st.write("Columns after renaming:", tc_df.columns)

      # Step 2: Validate required columns
      required_columns = [
          config["date_column"],
          config["time_column"],
          config["advt_theme_column"]
      ]
      missing_columns = [col for col in required_columns if col not in tc_df.columns]
      if missing_columns:
          raise KeyError(f"Missing required columns in TC sheet: {missing_columns}")

      # Step 3: Apply transformations based on the configuration
      tc_df[config["date_column"]] = pd.to_datetime(tc_df[config["date_column"]], dayfirst=True)
      tc_df['Dd'] = tc_df[config["date_column"]].dt.day
      tc_df['Mn'] = tc_df[config["date_column"]].dt.month
      tc_df['Yr'] = tc_df[config["date_column"]].dt.year

      # Normalize case if specified
      if config["additional_transformations"]["normalize_case"]:
          tc_df['Advt_Theme'] = tc_df[config["advt_theme_column"]].str.lower()
      else:
          tc_df['Advt_Theme'] = tc_df[config["advt_theme_column"]]

      # Extract start and end times from the time belt
      def parse_time_belt(time_belt):
          """Parse the time belt into start and end times."""
          try:
              start_time, end_time = time_belt.split('-')
              start_time = pd.to_datetime(start_time.strip(), format='%H:%M').time()
              end_time = pd.to_datetime(end_time.strip(), format='%H:%M').time()
              return start_time, end_time
          except ValueError:
              st.warning(f"Invalid time belt format: {time_belt}. Skipping row.")
              return None, None

      tc_df[['Start_Time', 'End_Time']] = tc_df[config["time_column"]].apply(
          lambda x: pd.Series(parse_time_belt(x))
      )

      # Drop rows with invalid time belts
      tc_df = tc_df.dropna(subset=['Start_Time', 'End_Time'])

      return tc_df[['Advt_Theme', 'Start_Time', 'End_Time', 'Dd', 'Mn', 'Yr']]
  except KeyError as e:
      st.error(str(e))
      return pd.DataFrame()

def filter_by_channel(df, channel, config):
 """Filter LMRB data by selected channel and rename columns if specified."""
 # Step 1: Rename columns if specified in the configuration
 if "rename_columns" in config and config["rename_columns"]:
     df = df.rename(columns=config["rename_columns"])

 # Step 2: Filter by channel
 return df[df["Channel"] == channel]

def match_tc_with_lmrb_by_date(tc_df, lmrb_df, config):
  """
  Match TC data with LMRB data based on the full date (day, month, year).
  """
  if tc_df.empty or lmrb_df.empty:
      st.error("One or both DataFrames are empty.")
      return pd.DataFrame()

  # Step 1: Create a Full_Date column in both TC and LMRB DataFrames
  tc_df['Full_Date'] = tc_df.apply(lambda row: f"{row['Yr']}-{row['Mn']:02d}-{row['Dd']:02d}", axis=1)
  tc_df['Full_Date'] = pd.to_datetime(tc_df['Full_Date'], format='%Y-%m-%d')

  lmrb_df['Full_Date'] = lmrb_df.apply(lambda row: f"{row['Yr']}-{row['Mn']:02d}-{row['Dd']:02d}", axis=1)
  lmrb_df['Full_Date'] = pd.to_datetime(lmrb_df['Full_Date'], format='%Y-%m-%d')

  matched_rows = []

  # Step 2: Iterate through each row in the TC DataFrame
  for _, tc_row in tc_df.iterrows():
      # Filter LMRB rows where the Full_Date matches the TC row's Full_Date
      date_matched_lmrb = lmrb_df[lmrb_df['Full_Date'] == tc_row['Full_Date']]

      # Append the matched rows to the list
      matched_rows.append(date_matched_lmrb)

  # Combine all matched rows into a single DataFrame
  matched_df = pd.concat(matched_rows) if matched_rows else pd.DataFrame()
  return matched_df

def highlight_matched_rows_excel(lmrb_df, matched_indices, highlight_color):
 """Highlight matched rows in the LMRB DataFrame when exporting to Excel."""
 # Add a 'Matched' column to indicate matched rows
 lmrb_df['Matched'] = False  # Initialize all rows as not matched
 lmrb_df.loc[matched_indices, 'Matched'] = True  # Mark matched rows as True

 # Create an Excel workbook
 wb = Workbook()
 ws = wb.active

 # Write headers
 for col_num, column_name in enumerate(lmrb_df.columns, start=1):
     ws.cell(row=1, column=col_num, value=column_name)

 # Write data with conditional formatting
 for row_num, row_data in enumerate(lmrb_df.itertuples(index=False), start=2):
     for col_num, cell_value in enumerate(row_data, start=1):
         cell = ws.cell(row=row_num, column=col_num, value=cell_value)
         if col_num == len(row_data):  # Check the 'Matched' column
             if row_data[-1]:  # If 'Matched' is True
                 for cell_to_highlight in ws[row_num]:
                     cell_to_highlight.fill = PatternFill(start_color=highlight_color.lstrip("#"), end_color=highlight_color.lstrip("#"), fill_type="solid")

 return wb

def main():
 # Set page configuration
 st.set_page_config(page_title="Neth FM Data Processor", layout="wide")

 # Sidebar for file uploads
 st.sidebar.title("Upload Files")
 uploaded_lmrb_file = st.sidebar.file_uploader("Upload LMRB Data Sheet", type=["xlsx", "xls", "csv"])
 uploaded_tc_file = st.sidebar.file_uploader("Upload TC Sheet", type=["xlsx", "xls", "csv"])

 # Main title
 st.title("Neth FM Data Processor")
 st.write("Process Neth FM data efficiently by uploading sheets, selecting channels, and performing transformations.")

 if uploaded_lmrb_file:
     try:
         # Progress bar for file processing
         with st.spinner("Processing LMRB Data Sheet..."):
             # Read the uploaded file
             if uploaded_lmrb_file.name.endswith(".xlsx"):
                 lmrb_df = pd.read_excel(uploaded_lmrb_file, engine='openpyxl')
             elif uploaded_lmrb_file.name.endswith(".xls"):
                 lmrb_df = pd.read_excel(uploaded_lmrb_file, engine='xlrd')
             else:
                 lmrb_df = pd.read_csv(uploaded_lmrb_file)

         # Select Channel
         if "Channel" in lmrb_df.columns:
             channel = st.selectbox("Select Channel", options=lmrb_df["Channel"].unique())
             
             # Load channel-specific configuration
             channel_config = load_channel_config(channel)
             if not channel_config:
                 st.error("Failed to load channel-specific configuration.")
                 return
             
             # Filter LMRB data and rename columns dynamically
             filtered_df = filter_by_channel(lmrb_df, channel, channel_config)
             st.success(f"Filtered Data for Channel: {channel}")
             st.dataframe(filtered_df)

             if uploaded_tc_file:
                 with st.spinner("Processing TC Sheet..."):
                     if uploaded_tc_file.name.endswith(".xlsx"):
                         tc_df = pd.read_excel(uploaded_tc_file, engine='openpyxl')
                     elif uploaded_tc_file.name.endswith(".xls"):
                         tc_df = pd.read_excel(uploaded_tc_file, engine='xlrd')
                     else:
                         tc_df = pd.read_csv(uploaded_tc_file)

                     # Transform TC data based on channel-specific configuration
                     transformed_tc_df = transform_tc_data(tc_df, channel_config)
                     st.success("Transformed TC Data:")
                     st.dataframe(transformed_tc_df)

                     # Match TC data with cleaned LMRB data
                     matched_df, matched_indices = match_tc_with_lmrb_by_date(transformed_tc_df, filtered_df, channel_config)
                     st.success("Matched Data:")
                     st.dataframe(matched_df)

                     total_matched_count = len(matched_df)
                     st.write(f"Total Matched Data: {total_matched_count}")

                     # Choose highlight color
                     highlight_color = st.color_picker("Choose Highlight Color", "#FFFF00")

                     # Highlight matched rows in cleaned LMRB and export to Excel
                     highlighted_workbook = highlight_matched_rows_excel(filtered_df, matched_indices, highlight_color)

                     # Save the workbook to a BytesIO object
                     output_highlighted = BytesIO()
                     highlighted_workbook.save(output_highlighted)
                     output_highlighted.seek(0)

                     # Buttons for downloading highlighted LMRB data and matched data
                     col1, col2 = st.columns(2)
                     with col1:
                         st.download_button(
                             label="Download Highlighted LMRB Sheet",
                             data=output_highlighted,
                             file_name="highlighted_lmrb.xlsx",
                             key="highlighted_lmrb_download",
                             help="Download the highlighted LMRB data sheet."
                         )
                     with col2:
                         matched_output = BytesIO()
                         matched_df.to_excel(matched_output, index=False, engine='openpyxl')
                         matched_output.seek(0)
                         st.download_button(
                             label="Download Matched Data",
                             data=matched_output,
                             file_name="matched_data.xlsx",
                             key="matched_data_download",
                             help="Download the matched data sheet."
                         )
         else:
             st.error("The uploaded LMRB file does not contain a 'Channel' column.")
     except Exception as e:
         st.error(f"Error processing the uploaded LMRB file: {e}")
 else:
     st.info("Please upload an LMRB data sheet to begin.")

if __name__ == "__main__":
 main()